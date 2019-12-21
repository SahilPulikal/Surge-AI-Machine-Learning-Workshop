import openpyxl

filepath = "./Assignments/Assignment 1/1.2 output.xlsx"
wb = openpyxl.Workbook()

wb.save(filepath)

# Load the xlsx file by using path or name
wb = openpyxl.load_workbook("./Assignments/Assignment 1/1.2 output.xlsx")

# Get workbook active sheet
sheet = wb.active
# Iterate through rows
for i in range(1,11):
    # Iterate through columns
    for j in range(1,27):
        # Assign value to specific cell
        sheet.cell(row=i,column=j).value=chr(64+j)+str(i)
# Save the xlsx file        
wb.save('./Assignments/Assignment 1/1.2 output.xlsx') 